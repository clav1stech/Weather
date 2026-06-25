import os
import sys
import glob
import datetime
import re
import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from functools import reduce
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# --- Configuration ---
DOSSIER_SORTIE = "Forecasts"
os.makedirs(DOSSIER_SORTIE, exist_ok=True)

# Switch du run souhaité : "0Z" ou "12Z"
# Priorité : argument CLI (ex: `python Forecast.py 12Z`), puis variable d'env, sinon "0Z".
_run_arg = sys.argv[1] if len(sys.argv) > 1 and sys.argv[1] in ("0Z", "12Z") else None
CHOIX_RUN = _run_arg or os.environ.get("FORECAST_RUN", "0Z")
run_param = "0" if CHOIX_RUN == "0Z" else "12"
heure_attendue = 12 if CHOIX_RUN == "12Z" else 0

modeles = [
    {"nom": "ECMWF ENS", "sheet": "ECMWF", "url": f"https://www.meteociel.fr/modeles/ecmwfens_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&run={run_param}"},
    {"nom": "AIFS ENS", "sheet": "AIFS", "url": f"https://www.meteociel.fr/modeles/ecmwfens_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&aifs=1&run={run_param}"},
    {"nom": "GEFS", "sheet": "GEFS", "url": f"https://www.meteociel.fr/modeles/gefs_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&run={run_param}"}
]

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# --- Fonctions Utilitaires ---
def parse_file_datetime(filepath):
    match = re.search(r'Forecast-(\d{8})(?:-(0Z|12Z))?\.xlsx', os.path.basename(filepath))
    if match:
        dt = datetime.datetime.strptime(match.group(1), "%d%m%Y")
        run = match.group(2) if match.group(2) else "12Z" 
        heures = 12 if run == "12Z" else 0
        return dt + datetime.timedelta(hours=heures)
    return datetime.datetime.min

def parser_tableau_meteociel(html):
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find('table', class_='gefs')
    if not table: return None, None
    
    data, colors = [], []
    for tr in table.find_all('tr'):
        row_data, row_colors = [], []
        for td in tr.find_all(['td', 'th']):
            row_data.append(td.get_text(strip=True))
            bg = td.get('bgcolor')
            row_colors.append(bg.lstrip('#') if bg else None)
        data.append(row_data)
        colors.append(row_colors)
        
    df = pd.DataFrame(data[1:], columns=data[0])
    for col in df.columns:
        if col not in ['Date', 'Ech.']:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            
    return df, colors[1:]

def appliquer_style_tableau(ws, start_row, style_type="standard", df_cols=None):
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_white = Font(color="FFFFFF", bold=True)
    
    default_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    color_ecmwf = PatternFill(start_color="1F618D", end_color="1F618D", fill_type="solid") 
    color_aifs = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")  
    color_gefs = PatternFill(start_color="B9770E", end_color="B9770E", fill_type="solid")  
    color_div = PatternFill(start_color="7D3C98", end_color="7D3C98", fill_type="solid")   

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15 if col == 1 else 11
        
        header_fill = default_fill
        if df_cols:
            col_name = df_cols[col - 1]
            if "ECMWF" in col_name: header_fill = color_ecmwf
            elif "AIFS" in col_name: header_fill = color_aifs
            elif "GEFS" in col_name: header_fill = color_gefs
            elif style_type == "compare" and "Divergence" in col_name: header_fill = color_div

        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col, max_col=col):
            cell = row[0]
            cell.alignment = center_align
            cell.border = thin_border
            if cell.row == start_row:
                cell.fill = header_fill
                cell.font = font_white

# --- Initialisation ---
wb = Workbook()
wb.remove(wb.active)

model_dataframes = {}
model_colors = {}
runs_info = []

# --- 1. Extraction Tolérante ---
extractions_brutes = []

for mod in modeles:
    print(f"Téléchargement de {mod['nom']}...")
    try:
        response = requests.get(mod["url"], headers=headers, timeout=10)
        response.encoding = 'windows-1252'
        html = response.text
        
        match = re.search(r'(Run .*?(\d{2}/\d{2}/\d{4})\s+(\d{1,2})[Zz])', html)
        
        if not match:
            print(f"⚠️ AVERTISSEMENT : Impossible de lire la date pour {mod['nom']} (Tableau non publié ou format cassé).")
            continue
            
        run_info_complet = match.group(1)
        run_date_ext = match.group(2)
        run_heure_ext = int(match.group(3))
        
        df, row_colors = parser_tableau_meteociel(html)
        if df is not None:
            extractions_brutes.append({
                'nom': mod['nom'],
                'sheet': mod['sheet'],
                'df': df,
                'colors': row_colors,
                'date': run_date_ext,
                'heure': run_heure_ext,
                'info': run_info_complet
            })
    except Exception as e:
        print(f"⚠️ AVERTISSEMENT : Échec de la connexion pour {mod['nom']} ({e}).")

if not extractions_brutes:
    sys.exit("\n❌ ERREUR CRITIQUE : Absolument aucun modèle n'a pu être récupéré. Exécution annulée.")

# --- 2. Tri et Vote Majoritaire (Gestion de la désynchronisation) ---
# Étape A : On s'assure d'abord que les modèles ont bien respecté l'heure demandée via l'URL
modeles_heure_valide = []
for m in extractions_brutes:
    if m['heure'] == heure_attendue:
        modeles_heure_valide.append(m)
    else:
        print(f"  ❌ {m['nom']} écarté (Mauvais run : {m['heure']}Z au lieu de {heure_attendue}Z)")

if not modeles_heure_valide:
    sys.exit(f"\n❌ ERREUR : Aucun modèle n'est disponible pour le run {CHOIX_RUN}.")

# Étape B : Vote majoritaire sur la DATE du jour (pour isoler les modèles en retard)
dates = [m['date'] for m in modeles_heure_valide]
counts = Counter(dates)

# En cas d'égalité stricte (ex: 1 modèle à J, 1 modèle à J-1), on privilégie la date la plus récente.
def rank_date(item):
    date_str, count = item
    return (count, datetime.datetime.strptime(date_str, "%d/%m/%Y"))

majority_date = sorted(counts.items(), key=rank_date, reverse=True)[0][0]

date_file_str = datetime.datetime.strptime(majority_date, "%d/%m/%Y").strftime("%d%m%Y")
fichier_actuel = os.path.join(DOSSIER_SORTIE, f"Forecast-{date_file_str}-{CHOIX_RUN}.xlsx")

print(f"\n👉 Date majoritaire retenue : {majority_date} (Run {CHOIX_RUN})")

for m in modeles_heure_valide:
    if m['date'] == majority_date:
        print(f"  ✅ {m['nom']} intégré au super-ensemble.")
        model_dataframes[m['nom']] = m['df']
        model_colors[m['nom']] = m['colors']
        runs_info.append(f"{m['nom']} ({m['info']})")
        
        # Création de sa feuille individuelle
        ws_data = wb.create_sheet(title=m["sheet"])
        ws_data["A1"], ws_data["A2"] = f"Modèle : {m['nom']}", f"Informations : {m['info']}"
        ws_data["A1"].font = ws_data["A2"].font = Font(bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(m['df'], index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 4: 
                    color_hex = m['colors'][r_idx - 5][c_idx - 1]
                    if color_hex:
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        appliquer_style_tableau(ws_data, start_row=4)
    else:
        print(f"  ❌ {m['nom']} écarté (En retard : toujours bloqué à la date du {m['date']})")

# --- 3. Définition de l'historique ---
fichiers_existants = glob.glob(os.path.join(DOSSIER_SORTIE, "Forecast-*.xlsx"))
fichiers_tries = sorted(fichiers_existants, key=parse_file_datetime, reverse=True)

current_dt = parse_file_datetime(fichier_actuel)
limite_dt = current_dt - datetime.timedelta(days=5)

fichiers_precedents = [
    f for f in fichiers_tries 
    if parse_file_datetime(f) < current_dt and parse_file_datetime(f) >= limite_dt
][:10]

# --- 4. Synthèses ---
print("\nCalcul des statistiques de synthèse...")

dfs_to_merge, all_members = [], []
dfs_compare_list = []
det_col_names = {}  # short_nom -> nom de colonne renommée dans df_super

# Noms possibles de la colonne déterministe selon le modèle :
# ECMWF/AIFS → "DET", GEFS → "GFS" (run haute résolution, distinct des membres d'ensemble).
DET_RAW_NAMES = {'DET', 'GFS'}

for mod_nom, df_mod in model_dataframes.items():
    short_nom = mod_nom.replace(" ENS", "")
    # Détection tolérante du déterministe (gère le "GFS" du GEFS, sinon il serait
    # compté à tort comme un membre d'ensemble et fausserait les statistiques).
    det_src = next((c for c in df_mod.columns
                    if str(c).strip().upper() in DET_RAW_NAMES), None)
    has_det = det_src is not None
    exclude = ['Date', 'Ech.'] + ([det_src] if has_det else [])
    membres = [c for c in df_mod.columns if c not in exclude]

    extra_cols = [det_src] if has_det else []
    df_temp = df_mod[['Date', 'Ech.'] + membres + extra_cols].copy()
    new_cols = [f"{short_nom}_{c}" for c in membres]
    det_renamed = [f"{short_nom}_DET"] if has_det else []
    df_temp.columns = ['Date', 'Ech.'] + new_cols + det_renamed
    all_members.extend(new_cols)
    if has_det:
        det_col_names[short_nom] = f"{short_nom}_DET"
    dfs_to_merge.append(df_temp)

    m_indiv = df_mod[membres].apply(pd.to_numeric, errors='coerce')
    df_mod_stats = pd.DataFrame()
    df_mod_stats['Date'] = df_mod['Date']
    df_mod_stats['Ech'] = df_mod['Ech.']
    df_mod_stats[f'{short_nom} P10'] = m_indiv.quantile(0.10, axis=1).round(1)
    df_mod_stats[f'{short_nom} Médiane'] = m_indiv.median(axis=1).round(1)
    if has_det:
        df_mod_stats[f'{short_nom} DET'] = pd.to_numeric(df_mod[det_src], errors='coerce').round(1)
    df_mod_stats[f'{short_nom} P90'] = m_indiv.quantile(0.90, axis=1).round(1)
    df_mod_stats[f'{short_nom} Spread'] = (df_mod_stats[f'{short_nom} P90'] - df_mod_stats[f'{short_nom} P10']).round(1)
    dfs_compare_list.append(df_mod_stats)

if dfs_to_merge:
    # --- Super-Ensemble ---
    df_super = reduce(lambda left, right: pd.merge(left, right, on=['Date', 'Ech.'], how='outer'), dfs_to_merge)
    df_super = df_super[df_super['Date'].str.contains(' 12Z', na=False)].copy()
    
    df_super['Date'] = df_super['Date'].str.split(' ').str[0]
    df_super['Date'] = pd.to_datetime(df_super['Date'], errors='coerce').dt.date
    
    df_super['Ech_num'] = pd.to_numeric(df_super['Ech.'], errors='coerce')
    df_super = df_super.sort_values('Ech_num').drop(columns=['Ech_num']).reset_index(drop=True)
    
    m = df_super[all_members]
    
    df_stats = pd.DataFrame()
    df_stats['Date'] = df_super['Date']
    df_stats['Ech'] = df_super['Ech.']
    df_stats['Min'] = m.min(axis=1).round(1)
    df_stats['P10'] = m.quantile(0.10, axis=1).round(1)
    df_stats['P25'] = m.quantile(0.25, axis=1).round(1)
    df_stats['Médiane'] = m.median(axis=1).round(1)
    for short_nom, det_col in det_col_names.items():
        if det_col in df_super.columns:
            df_stats[f'{short_nom} DET'] = pd.to_numeric(df_super[det_col], errors='coerce').round(1)
    df_stats['P75'] = m.quantile(0.75, axis=1).round(1)
    df_stats['P90'] = m.quantile(0.90, axis=1).round(1)
    df_stats['Max'] = m.max(axis=1).round(1)
    df_stats['Spread'] = (df_stats['P90'] - df_stats['P10']).round(1)
    df_stats['Ecart-type'] = m.std(axis=1).round(2)
    df_stats['Proba > 20°'] = ((m > 20).sum(axis=1) / m.notna().sum(axis=1)).fillna(0)
    
    df_stats_ech_num = pd.to_numeric(df_stats['Ech'], errors='coerce')
    
    # --- Deltas ---
    for prev_file in fichiers_precedents:
        try:
            prev_dt = parse_file_datetime(prev_file)
            delta_hours = int((current_dt - prev_dt).total_seconds() / 3600)
            
            df_prev = pd.read_excel(prev_file, sheet_name="Synthèse", skiprows=4)
            if 'Médiane' in df_prev.columns and 'Ech' in df_prev.columns:
                
                df_prev['Ech_Num'] = pd.to_numeric(df_prev['Ech'], errors='coerce')
                prev_med_map = df_prev.set_index('Ech_Num')['Médiane'].to_dict()
                
                heure_str = "12Z" if prev_dt.hour == 12 else "0Z"
                col_delta = f"Δ {prev_dt.strftime('%d-%b')} {heure_str}"
                
                mediane_prev_aligned = df_stats_ech_num.apply(lambda e: prev_med_map.get(e + delta_hours, np.nan))
                df_stats[col_delta] = (df_stats['Médiane'] - mediane_prev_aligned).round(1)
        except Exception:
            pass 
            
    # --- Synthèse Comparée (Individuelle) ---
    df_compare = reduce(lambda left, right: pd.merge(left, right, on=['Date', 'Ech'], how='outer'), dfs_compare_list)
    df_compare = df_compare[df_compare['Date'].str.contains(' 12Z', na=False)].copy()
    
    df_compare['Date'] = df_compare['Date'].str.split(' ').str[0]
    df_compare['Date'] = pd.to_datetime(df_compare['Date'], errors='coerce').dt.date
    
    df_compare['Ech_num'] = pd.to_numeric(df_compare['Ech'], errors='coerce')
    df_compare = df_compare.sort_values('Ech_num').drop(columns=['Ech_num']).reset_index(drop=True)
    
    # Divergence (sécurisée même s'il ne reste qu'un seul modèle)
    median_cols = [c for c in df_compare.columns if 'Médiane' in c]
    if len(median_cols) > 1:
        df_compare['Divergence Modèles'] = (df_compare[median_cols].max(axis=1) - df_compare[median_cols].min(axis=1)).round(1)
    else:
        df_compare['Divergence Modèles'] = np.nan
    
    # --- Écriture Feuille 1 : Synthèse Comparée ---
    ws_comp = wb.create_sheet(title="Comparaison Modèles", index=0)
    ws_comp["A1"] = "Comparaison des Tendances par Modèle - Echéances 12Z"
    ws_comp["A2"] = f"Généré le : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_comp["A3"] = f"Divergence : Indique l'écart de température entre le modèle le plus chaud et le plus froid."
    ws_comp["A1"].font = ws_comp["A2"].font = ws_comp["A3"].font = Font(bold=True)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_compare, index=False, header=True), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_comp.cell(row=r_idx, column=c_idx, value=value)
            col_name = df_compare.columns[c_idx-1]
            
            if col_name == 'Date' and r_idx > 5:
                cell.number_format = 'dd-mmm'
                
            if col_name == 'Divergence Modèles' and r_idx > 5 and pd.notna(value):
                if value >= 4.0:
                    cell.font = Font(color="D32F2F", bold=True)
                elif value <= 1.5:
                    cell.font = Font(color="1976D2", bold=True)
                    
    appliquer_style_tableau(ws_comp, start_row=5, style_type="compare", df_cols=df_compare.columns.tolist())
    ws_comp.freeze_panes = "C6"

    # --- Écriture Feuille 2 : Synthèse Globale ---
    ws_synth = wb.create_sheet(title="Synthèse", index=1)
    ws_synth["A1"] = "Synthèse Multi-Modèles (Super-Ensemble) - Echéances 12Z"
    ws_synth["A2"] = f"Généré le : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_synth["A3"] = f"Runs inclus : { ' | '.join(runs_info) }"
    ws_synth["A1"].font = ws_synth["A2"].font = ws_synth["A3"].font = Font(bold=True)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_stats, index=False, header=True), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_synth.cell(row=r_idx, column=c_idx, value=value)
            col_name = df_stats.columns[c_idx-1]
            
            if col_name == 'Date' and r_idx > 5:
                cell.number_format = 'dd-mmm'
            
            if col_name == 'Proba > 20°' and r_idx > 5:
                cell.number_format = '0.0%'
                
            if col_name.startswith('Δ') and r_idx > 5 and pd.notna(value):
                if value > 0:
                    cell.font = Font(color="D32F2F", bold=True)
                elif value < 0:
                    cell.font = Font(color="1976D2", bold=True)
                    
    appliquer_style_tableau(ws_synth, start_row=5, df_cols=df_stats.columns.tolist())
    ws_synth.freeze_panes = "C6"

# --- Sauvegarde ---
wb.save(fichier_actuel)
print(f"\n✅ Terminé ! Fichier sauvegardé sous : {fichier_actuel}")