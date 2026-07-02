import os
import sys
import glob
import datetime
import re
import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from functools import reduce
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# --- Configuration ---
DOSSIER_SORTIE = "legacy"
os.makedirs(DOSSIER_SORTIE, exist_ok=True)

# Switch du run souhaité : "0Z", "6Z", "12Z" ou "18Z".
# Priorité : argument CLI (ex: `python Forecast_legacy.py 12Z`), puis variable d'env, sinon "0Z".
# 0Z/12Z : run principal automatique (les 3 modèles, cf. workflow schedule).
# 6Z/18Z : run manuel only (workflow_dispatch) — ECMWF ENS n'y est pas publié en
# intégralité par Météociel (cf. config.EXPECTED_CYCLES_BY_LABEL), on ne scrape
# donc que AIFS/GEFS, qui tournent réellement 4×/j.
RUN_HEURES = {"0Z": 0, "6Z": 6, "12Z": 12, "18Z": 18}
_run_arg = sys.argv[1] if len(sys.argv) > 1 and sys.argv[1] in RUN_HEURES else None
CHOIX_RUN = _run_arg or os.environ.get("FORECAST_RUN", "0Z")
heure_attendue = RUN_HEURES[CHOIX_RUN]
run_param = str(heure_attendue)

modeles = [
    {"nom": "ECMWF ENS", "sheet": "ECMWF", "url": f"https://www.meteociel.fr/modeles/ecmwfens_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&run={run_param}"},
    {"nom": "AIFS ENS", "sheet": "AIFS", "url": f"https://www.meteociel.fr/modeles/ecmwfens_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&aifs=1&run={run_param}"},
    {"nom": "GEFS", "sheet": "GEFS", "url": f"https://www.meteociel.fr/modeles/gefs_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&run={run_param}"}
]
if CHOIX_RUN in ("6Z", "18Z"):
    modeles = [m for m in modeles if m["nom"] != "ECMWF ENS"]

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# Ordre canonique des modèles (ordre des feuilles + priorité d'affichage).
ORDRE_MODELES = [m["nom"] for m in modeles]
SHEET_PAR_NOM = {m["nom"]: m["sheet"] for m in modeles}
NOM_PAR_SHEET = {m["sheet"]: m["nom"] for m in modeles}

# Échéance maximale nominale par modèle (en heures). Sert à repérer les runs
# publiés partiellement : météociel ne met d'abord en ligne qu'une partie des
# échéances. Un modèle dont l'échéance max est nettement sous son nominal est
# considéré « pas encore prêt » et écarté — sinon il tronquerait les graphiques
# et risquerait d'écraser un run complet déjà en stock.
ECHEANCE_MAX_NOMINALE = {"ECMWF ENS": 360, "AIFS ENS": 360, "GEFS": 384}
SEUIL_COMPLETUDE = 0.9  # fraction du nominal requise pour juger un run « complet »

# --- Fonctions Utilitaires ---
def parse_file_datetime(filepath):
    match = re.search(r'Forecast-(\d{8})(?:-(0Z|6Z|12Z|18Z))?\.xlsx', os.path.basename(filepath))
    if match:
        dt = datetime.datetime.strptime(match.group(1), "%d%m%Y")
        run = match.group(2) if match.group(2) else "12Z"
        return dt + datetime.timedelta(hours=RUN_HEURES[run])
    return datetime.datetime.min

def parser_tableau_meteociel(html):
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find('table', class_='gefs')
    if not table: return None, None
    
    data, colors = [], []
    for tr in table.find_all('tr'):
        row_data, row_colors = [], []
        for td in tr.find_all(['td', 'th']):
            row_data.append(td.get_text(strip=True))
            bg = td.get('bgcolor')
            row_colors.append(bg.lstrip('#') if bg else None)
        data.append(row_data)
        colors.append(row_colors)
        
    df = pd.DataFrame(data[1:], columns=data[0])
    for col in df.columns:
        if col not in ['Date', 'Ech.']:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            
    return df, colors[1:]

def appliquer_style_tableau(ws, start_row, style_type="standard", df_cols=None):
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_white = Font(color="FFFFFF", bold=True)
    
    default_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    color_ecmwf = PatternFill(start_color="1F618D", end_color="1F618D", fill_type="solid") 
    color_aifs = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")  
    color_gefs = PatternFill(start_color="B9770E", end_color="B9770E", fill_type="solid")  
    color_div = PatternFill(start_color="7D3C98", end_color="7D3C98", fill_type="solid")   

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15 if col == 1 else 11
        
        header_fill = default_fill
        if df_cols:
            col_name = df_cols[col - 1]
            if "ECMWF" in col_name: header_fill = color_ecmwf
            elif "AIFS" in col_name: header_fill = color_aifs
            elif "GEFS" in col_name: header_fill = color_gefs
            elif style_type == "compare" and "Divergence" in col_name: header_fill = color_div

        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col, max_col=col):
            cell = row[0]
            cell.alignment = center_align
            cell.border = thin_border
            if cell.row == start_row:
                cell.fill = header_fill
                cell.font = font_white

def max_echeance(df):
    """Échéance maximale (en heures) présente dans un tableau modèle."""
    return pd.to_numeric(df['Ech.'], errors='coerce').max()


def lire_feuille_modele(fichier, sheet):
    """Relit une feuille modèle déjà stockée (valeurs + couleurs des cellules).

    Permet de récupérer un modèle déjà en stock pour compléter un run partiel
    sans le re-télécharger (météociel l'ayant souvent déjà fait passer au run
    suivant). Renvoie {'df', 'colors', 'info'} ou None si la feuille est absente.
    """
    if not os.path.exists(fichier):
        return None
    try:
        wb_src = load_workbook(fichier, data_only=True)
    except Exception as e:
        print(f"  ⚠️ Impossible de relire {os.path.basename(fichier)} ({e}).")
        return None
    if sheet not in wb_src.sheetnames:
        return None

    ws = wb_src[sheet]
    info = ws['A2'].value or ''
    if isinstance(info, str) and info.startswith('Informations : '):
        info = info[len('Informations : '):]

    # En-tête en ligne 4, données à partir de la ligne 5 (cf. écriture ci-dessous).
    rows = list(ws.iter_rows(min_row=4))
    if not rows:
        return None
    header = [c.value for c in rows[0]]
    ncol = len(header)

    data, colors = [], []
    for r in rows[1:]:
        cells = list(r)[:ncol]
        vals = [c.value for c in cells]
        if all(v is None for v in vals):
            continue
        data.append(vals)
        ligne_couleurs = []
        for c in cells:
            hexa = None
            fill = c.fill
            if fill is not None and fill.patternType and fill.fgColor is not None:
                rgb = fill.fgColor.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    hexa = rgb[-6:]
            ligne_couleurs.append(hexa)
        colors.append(ligne_couleurs)

    df = pd.DataFrame(data, columns=header)
    for col in df.columns:
        if col not in ['Date', 'Ech.']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    return {'df': df, 'colors': colors, 'info': info}


def construire_et_sauver(model_dataframes, model_colors, model_infos, fichier_cible):
    """Construit le classeur (feuilles modèles + synthèses + deltas) et l'enregistre.

    `model_dataframes` / `model_colors` / `model_infos` sont indexés par nom de
    modèle et déjà ordonnés (ordre canonique). Toutes les statistiques sont
    recalculées à partir de l'union des modèles fournis.
    """
    wb = Workbook()
    wb.remove(wb.active)

    runs_info = [f"{nom} ({model_infos[nom]})" for nom in model_dataframes]

    # --- Feuilles individuelles par modèle ---
    for nom, df_mod in model_dataframes.items():
        sheet = SHEET_PAR_NOM[nom]
        colors = model_colors[nom]
        ws_data = wb.create_sheet(title=sheet)
        ws_data["A1"], ws_data["A2"] = f"Modèle : {nom}", f"Informations : {model_infos[nom]}"
        ws_data["A1"].font = ws_data["A2"].font = Font(bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(df_mod, index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 4:
                    ci, cj = r_idx - 5, c_idx - 1
                    color_hex = colors[ci][cj] if (ci < len(colors) and cj < len(colors[ci])) else None
                    if color_hex:
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        appliquer_style_tableau(ws_data, start_row=4)

    # --- Historique (runs strictement antérieurs, fenêtre de 5 jours) ---
    current_dt = parse_file_datetime(fichier_cible)
    limite_dt = current_dt - datetime.timedelta(days=5)
    fichiers_existants = glob.glob(os.path.join(DOSSIER_SORTIE, "Forecast-*.xlsx"))
    fichiers_tries = sorted(fichiers_existants, key=parse_file_datetime, reverse=True)
    fichiers_precedents = [
        f for f in fichiers_tries
        if parse_file_datetime(f) < current_dt and parse_file_datetime(f) >= limite_dt
    ][:10]

    # --- Synthèses ---
    dfs_to_merge, all_members = [], []
    dfs_compare_list = []
    det_col_names = {}  # short_nom -> nom de colonne renommée dans df_super

    # Noms possibles de la colonne déterministe selon le modèle :
    # ECMWF/AIFS → "DET", GEFS → "GFS" (run haute résolution, distinct des membres).
    DET_RAW_NAMES = {'DET', 'GFS'}

    for mod_nom, df_mod in model_dataframes.items():
        short_nom = mod_nom.replace(" ENS", "")
        det_src = next((c for c in df_mod.columns
                        if str(c).strip().upper() in DET_RAW_NAMES), None)
        has_det = det_src is not None
        exclude = ['Date', 'Ech.'] + ([det_src] if has_det else [])
        membres = [c for c in df_mod.columns if c not in exclude]

        extra_cols = [det_src] if has_det else []
        df_temp = df_mod[['Date', 'Ech.'] + membres + extra_cols].copy()
        new_cols = [f"{short_nom}_{c}" for c in membres]
        det_renamed = [f"{short_nom}_DET"] if has_det else []
        df_temp.columns = ['Date', 'Ech.'] + new_cols + det_renamed
        all_members.extend(new_cols)
        if has_det:
            det_col_names[short_nom] = f"{short_nom}_DET"
        dfs_to_merge.append(df_temp)

        m_indiv = df_mod[membres].apply(pd.to_numeric, errors='coerce')
        df_mod_stats = pd.DataFrame()
        df_mod_stats['Date'] = df_mod['Date']
        df_mod_stats['Ech'] = df_mod['Ech.']
        df_mod_stats[f'{short_nom} P10'] = m_indiv.quantile(0.10, axis=1).round(1)
        df_mod_stats[f'{short_nom} Médiane'] = m_indiv.median(axis=1).round(1)
        if has_det:
            df_mod_stats[f'{short_nom} DET'] = pd.to_numeric(df_mod[det_src], errors='coerce').round(1)
        df_mod_stats[f'{short_nom} P90'] = m_indiv.quantile(0.90, axis=1).round(1)
        df_mod_stats[f'{short_nom} Spread'] = (df_mod_stats[f'{short_nom} P90'] - df_mod_stats[f'{short_nom} P10']).round(1)
        dfs_compare_list.append(df_mod_stats)

    if dfs_to_merge:
        # --- Super-Ensemble ---
        df_super = reduce(lambda left, right: pd.merge(left, right, on=['Date', 'Ech.'], how='outer'), dfs_to_merge)
        df_super = df_super[df_super['Date'].str.contains(' 12Z', na=False)].copy()

        df_super['Date'] = df_super['Date'].str.split(' ').str[0]
        df_super['Date'] = pd.to_datetime(df_super['Date'], errors='coerce').dt.date

        df_super['Ech_num'] = pd.to_numeric(df_super['Ech.'], errors='coerce')
        df_super = df_super.sort_values('Ech_num').drop(columns=['Ech_num']).reset_index(drop=True)

        m = df_super[all_members]

        df_stats = pd.DataFrame()
        df_stats['Date'] = df_super['Date']
        df_stats['Ech'] = df_super['Ech.']
        df_stats['Min'] = m.min(axis=1).round(1)
        df_stats['P10'] = m.quantile(0.10, axis=1).round(1)
        df_stats['P25'] = m.quantile(0.25, axis=1).round(1)
        df_stats['Médiane'] = m.median(axis=1).round(1)
        for short_nom, det_col in det_col_names.items():
            if det_col in df_super.columns:
                df_stats[f'{short_nom} DET'] = pd.to_numeric(df_super[det_col], errors='coerce').round(1)
        df_stats['P75'] = m.quantile(0.75, axis=1).round(1)
        df_stats['P90'] = m.quantile(0.90, axis=1).round(1)
        df_stats['Max'] = m.max(axis=1).round(1)
        df_stats['Spread'] = (df_stats['P90'] - df_stats['P10']).round(1)
        df_stats['Ecart-type'] = m.std(axis=1).round(2)
        df_stats['Proba > 20°'] = ((m > 20).sum(axis=1) / m.notna().sum(axis=1)).fillna(0)

        df_stats_ech_num = pd.to_numeric(df_stats['Ech'], errors='coerce')

        # --- Deltas ---
        for prev_file in fichiers_precedents:
            try:
                prev_dt = parse_file_datetime(prev_file)
                delta_hours = int((current_dt - prev_dt).total_seconds() / 3600)

                df_prev = pd.read_excel(prev_file, sheet_name="Synthèse", skiprows=4)
                if 'Médiane' in df_prev.columns and 'Ech' in df_prev.columns:

                    df_prev['Ech_Num'] = pd.to_numeric(df_prev['Ech'], errors='coerce')
                    prev_med_map = df_prev.set_index('Ech_Num')['Médiane'].to_dict()

                    heure_str = "12Z" if prev_dt.hour == 12 else "0Z"
                    col_delta = f"Δ {prev_dt.strftime('%d-%b')} {heure_str}"

                    mediane_prev_aligned = df_stats_ech_num.apply(lambda e: prev_med_map.get(e + delta_hours, np.nan))
                    df_stats[col_delta] = (df_stats['Médiane'] - mediane_prev_aligned).round(1)
            except Exception:
                pass

        # --- Synthèse Comparée (Individuelle) ---
        df_compare = reduce(lambda left, right: pd.merge(left, right, on=['Date', 'Ech'], how='outer'), dfs_compare_list)
        df_compare = df_compare[df_compare['Date'].str.contains(' 12Z', na=False)].copy()

        df_compare['Date'] = df_compare['Date'].str.split(' ').str[0]
        df_compare['Date'] = pd.to_datetime(df_compare['Date'], errors='coerce').dt.date

        df_compare['Ech_num'] = pd.to_numeric(df_compare['Ech'], errors='coerce')
        df_compare = df_compare.sort_values('Ech_num').drop(columns=['Ech_num']).reset_index(drop=True)

        # Divergence (sécurisée même s'il ne reste qu'un seul modèle)
        median_cols = [c for c in df_compare.columns if 'Médiane' in c]
        if len(median_cols) > 1:
            df_compare['Divergence Modèles'] = (df_compare[median_cols].max(axis=1) - df_compare[median_cols].min(axis=1)).round(1)
        else:
            df_compare['Divergence Modèles'] = np.nan

        # --- Écriture Feuille 1 : Synthèse Comparée ---
        ws_comp = wb.create_sheet(title="Comparaison Modèles", index=0)
        ws_comp["A1"] = "Comparaison des Tendances par Modèle - Echéances 12Z"
        ws_comp["A2"] = f"Généré le : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_comp["A3"] = f"Divergence : Indique l'écart de température entre le modèle le plus chaud et le plus froid."
        ws_comp["A1"].font = ws_comp["A2"].font = ws_comp["A3"].font = Font(bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(df_compare, index=False, header=True), start=5):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_comp.cell(row=r_idx, column=c_idx, value=value)
                col_name = df_compare.columns[c_idx-1]

                if col_name == 'Date' and r_idx > 5:
                    cell.number_format = 'dd-mmm'

                if col_name == 'Divergence Modèles' and r_idx > 5 and pd.notna(value):
                    if value >= 4.0:
                        cell.font = Font(color="D32F2F", bold=True)
                    elif value <= 1.5:
                        cell.font = Font(color="1976D2", bold=True)

        appliquer_style_tableau(ws_comp, start_row=5, style_type="compare", df_cols=df_compare.columns.tolist())
        ws_comp.freeze_panes = "C6"

        # --- Écriture Feuille 2 : Synthèse Globale ---
        ws_synth = wb.create_sheet(title="Synthèse", index=1)
        ws_synth["A1"] = "Synthèse Multi-Modèles (Super-Ensemble) - Echéances 12Z"
        ws_synth["A2"] = f"Généré le : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_synth["A3"] = f"Runs inclus : { ' | '.join(runs_info) }"
        ws_synth["A1"].font = ws_synth["A2"].font = ws_synth["A3"].font = Font(bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(df_stats, index=False, header=True), start=5):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_synth.cell(row=r_idx, column=c_idx, value=value)
                col_name = df_stats.columns[c_idx-1]

                if col_name == 'Date' and r_idx > 5:
                    cell.number_format = 'dd-mmm'

                if col_name == 'Proba > 20°' and r_idx > 5:
                    cell.number_format = '0.0%'

                if col_name.startswith('Δ') and r_idx > 5 and pd.notna(value):
                    if value > 0:
                        cell.font = Font(color="D32F2F", bold=True)
                    elif value < 0:
                        cell.font = Font(color="1976D2", bold=True)

        appliquer_style_tableau(ws_synth, start_row=5, df_cols=df_stats.columns.tolist())
        ws_synth.freeze_panes = "C6"

    wb.save(fichier_cible)
    print(f"✅ Fichier sauvegardé sous : {fichier_cible}")


# --- 1. Extraction Tolérante ---
def extraire_modeles():
    """Télécharge et parse les trois modèles. Renvoie la liste des extractions brutes."""
    extractions_brutes = []
    for mod in modeles:
        print(f"Téléchargement de {mod['nom']}...")
        try:
            response = requests.get(mod["url"], headers=headers, timeout=10)
            response.encoding = 'windows-1252'
            html = response.text

            match = re.search(r'(Run .*?(\d{2}/\d{2}/\d{4})\s+(\d{1,2})[Zz])', html)

            if not match:
                print(f"⚠️ AVERTISSEMENT : Impossible de lire la date pour {mod['nom']} (Tableau non publié ou format cassé).")
                continue

            run_info_complet = match.group(1)
            run_date_ext = match.group(2)
            run_heure_ext = int(match.group(3))

            df, row_colors = parser_tableau_meteociel(html)
            if df is not None:
                extractions_brutes.append({
                    'nom': mod['nom'],
                    'sheet': mod['sheet'],
                    'df': df,
                    'colors': row_colors,
                    'date': run_date_ext,
                    'heure': run_heure_ext,
                    'info': run_info_complet
                })
        except Exception as e:
            print(f"⚠️ AVERTISSEMENT : Échec de la connexion pour {mod['nom']} ({e}).")
    return extractions_brutes


def router_et_construire(extractions_brutes):
    """Range chaque modèle dans le fichier de SA date de run, complète l'existant,
    écarte les runs aux échéances partielles, puis (re)construit chaque fichier."""
    if not extractions_brutes:
        sys.exit("\n❌ ERREUR CRITIQUE : Absolument aucun modèle n'a pu être récupéré. Exécution annulée.")

    # --- 2. Filtrage : heure demandée puis complétude des échéances ---
    # Étape A : on ne garde que les modèles ayant bien respecté l'heure demandée.
    modeles_heure_valide = []
    for m in extractions_brutes:
        if m['heure'] == heure_attendue:
            modeles_heure_valide.append(m)
        else:
            print(f"  ❌ {m['nom']} écarté (Mauvais run : {m['heure']}Z au lieu de {heure_attendue}Z)")

    if not modeles_heure_valide:
        sys.exit(f"\n❌ ERREUR : Aucun modèle n'est disponible pour le run {CHOIX_RUN}.")

    # Étape B : on écarte les modèles publiés partiellement (échéances tronquées).
    # Météociel ne met d'abord en ligne qu'une partie des échéances : intégrer un tel
    # modèle tronquerait les graphiques et pourrait écraser un run complet en stock.
    # On préfère attendre que les échéances soient (quasi) complètes.
    modeles_prets = []
    for m in modeles_heure_valide:
        nominal = ECHEANCE_MAX_NOMINALE.get(m['nom'])
        ech_max = max_echeance(m['df'])
        if nominal and (pd.isna(ech_max) or ech_max < SEUIL_COMPLETUDE * nominal):
            ech_txt = "—" if pd.isna(ech_max) else f"{int(ech_max)}h"
            print(f"  ⏳ {m['nom']} écarté (run encore partiel : échéances jusqu'à {ech_txt} "
                  f"/ {nominal}h attendues). La donnée existante est conservée.")
            continue
        modeles_prets.append(m)

    if not modeles_prets:
        sys.exit("\n⏳ Aucun modèle complet pour l'instant : rien n'a été écrit "
                 "(les runs précédents restent intacts).")

    # --- 3. Routage par date de run ---
    # Météociel fait passer les modèles au run suivant à des moments différents
    # (souvent GEFS, puis AIFS, puis ECMWF). Chaque modèle est donc rangé dans le
    # fichier correspondant à SA propre date de run, et non dans une date imposée à
    # tous. On complète/met à jour le fichier sans jamais perdre l'antériorité.
    groupes_par_date = {}
    for m in modeles_prets:
        groupes_par_date.setdefault(m['date'], []).append(m)

    print(f"\n👉 Run {CHOIX_RUN} — dates détectées : "
          + ", ".join(f"{d} ({', '.join(x['nom'] for x in g)})"
                      for d, g in sorted(groupes_par_date.items(),
                                         key=lambda kv: datetime.datetime.strptime(kv[0], '%d/%m/%Y'))))

    # On traite les dates de la plus ancienne à la plus récente : ainsi un fichier
    # récent peut calculer ses deltas sur un fichier antérieur déjà rafraîchi.
    for date_str in sorted(groupes_par_date,
                           key=lambda d: datetime.datetime.strptime(d, '%d/%m/%Y')):
        fresh = groupes_par_date[date_str]
        date_file_str = datetime.datetime.strptime(date_str, "%d/%m/%Y").strftime("%d%m%Y")
        fichier_cible = os.path.join(DOSSIER_SORTIE, f"Forecast-{date_file_str}-{CHOIX_RUN}.xlsx")

        print(f"\n📂 {os.path.basename(fichier_cible)} (run du {date_str})")

        # On part des modèles déjà stockés dans le fichier (pour compléter un run
        # partiel), puis on superpose les modèles fraîchement téléchargés.
        combined = {}  # nom -> {'df', 'colors', 'info'}
        if os.path.exists(fichier_cible):
            for sheet, nom in NOM_PAR_SHEET.items():
                existant = lire_feuille_modele(fichier_cible, sheet)
                if existant is not None:
                    combined[nom] = existant
                    print(f"  📦 {nom} déjà en stock (conservé).")

        for m in fresh:
            nom = m['nom']
            frais = {'df': m['df'], 'colors': m['colors'], 'info': m['info']}
            if nom in combined:
                # Garde-fou anti-régression : on ne remplace un modèle stocké que si
                # la version fraîche a au moins autant d'échéances (jamais complet → partiel).
                ech_frais = max_echeance(m['df'])
                ech_stock = max_echeance(combined[nom]['df'])
                if pd.notna(ech_stock) and pd.notna(ech_frais) and ech_frais < ech_stock:
                    print(f"  🛡️ {nom} : version stockée plus complète "
                          f"({int(ech_stock)}h > {int(ech_frais)}h) — conservée.")
                    continue
                print(f"  🔄 {nom} rafraîchi.")
            else:
                print(f"  ✅ {nom} ajouté.")
            combined[nom] = frais

        # Ordre canonique des feuilles (ECMWF, AIFS, GEFS).
        noms_ordonnes = [n for n in ORDRE_MODELES if n in combined]
        model_dataframes = {n: combined[n]['df'] for n in noms_ordonnes}
        model_colors = {n: combined[n]['colors'] for n in noms_ordonnes}
        model_infos = {n: combined[n]['info'] for n in noms_ordonnes}

        print(f"  → Modèles inclus : {', '.join(noms_ordonnes)}")
        construire_et_sauver(model_dataframes, model_colors, model_infos, fichier_cible)

    print("\n✅ Terminé !")


def tenter_rattrapage():
    """Après le run principal, tente de compléter le fichier du run opposé si des modèles y manquent.

    Ex : après le run 12Z, cherche le fichier 0Z le plus récent. Si ECMWF y manque
    encore, le télécharge et l'ajoute — uniquement si Météociel sert toujours la
    MÊME date et la MÊME heure de run que le fichier cible (double garde-fou).
    Ne remplace jamais un modèle déjà présent.

    Limité au cycle principal 0Z/12Z (cadence automatique) : les runs 6Z/18Z
    sont manuels, AIFS/GEFS seulement, et n'ont pas de mécanisme de rattrapage.
    """
    if CHOIX_RUN not in ("0Z", "12Z"):
        return
    autre_run = "0Z" if CHOIX_RUN == "12Z" else "12Z"
    autre_run_param = "0" if autre_run == "0Z" else "12"
    autre_heure = 0 if autre_run == "0Z" else 12

    pattern = os.path.join(DOSSIER_SORTIE, f"Forecast-*-{autre_run}.xlsx")
    fichiers = sorted(glob.glob(pattern), key=parse_file_datetime, reverse=True)
    if not fichiers:
        return

    fichier_cible = fichiers[0]
    date_attendue = parse_file_datetime(fichier_cible).date()

    try:
        wb_check = load_workbook(fichier_cible, read_only=True)
        sheets_presents = set(wb_check.sheetnames)
        wb_check.close()
    except Exception as e:
        print(f"  ⚠️ Rattrapage : impossible de lire {os.path.basename(fichier_cible)} ({e}).")
        return

    manquants = [m for m in modeles if SHEET_PAR_NOM[m['nom']] not in sheets_presents]
    if not manquants:
        return

    print(f"\n🔁 Rattrapage {autre_run} : {os.path.basename(fichier_cible)} "
          f"(date attendue : {date_attendue.strftime('%d/%m/%Y')})")
    print(f"   Modèles manquants : {', '.join(m['nom'] for m in manquants)}")

    urls_autre_run = {
        "ECMWF ENS": f"https://www.meteociel.fr/modeles/ecmwfens_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&run={autre_run_param}",
        "AIFS ENS":  f"https://www.meteociel.fr/modeles/ecmwfens_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&aifs=1&run={autre_run_param}",
        "GEFS":      f"https://www.meteociel.fr/modeles/gefs_table.php?ext=1&x=&lat=48.8621&lon=2.33936&ville=Paris&run={autre_run_param}",
    }

    nouveaux = {}
    for m_conf in manquants:
        nom = m_conf['nom']
        url = urls_autre_run[nom]
        print(f"   Téléchargement de {nom}...")
        try:
            response = requests.get(url, headers=headers, timeout=10)
            response.encoding = 'windows-1252'
            html = response.text

            match = re.search(r'(Run .*?(\d{2}/\d{2}/\d{4})\s+(\d{1,2})[Zz])', html)
            if not match:
                print(f"   ⚠️ {nom} : date illisible, ignoré.")
                continue

            run_date = datetime.datetime.strptime(match.group(2), "%d/%m/%Y").date()
            run_heure = int(match.group(3))

            if run_date != date_attendue:
                print(f"   ⚠️ {nom} : date {run_date.strftime('%d/%m/%Y')} ≠ "
                      f"{date_attendue.strftime('%d/%m/%Y')} (mauvais jour), ignoré.")
                continue
            if run_heure != autre_heure:
                print(f"   ⚠️ {nom} : sur le run {run_heure}Z (attendu {autre_run}), ignoré.")
                continue

            df, colors = parser_tableau_meteociel(html)
            if df is None:
                print(f"   ⚠️ {nom} : tableau non parseable, ignoré.")
                continue

            nominal = ECHEANCE_MAX_NOMINALE.get(nom)
            ech_max = max_echeance(df)
            if nominal and (pd.isna(ech_max) or ech_max < SEUIL_COMPLETUDE * nominal):
                ech_txt = "—" if pd.isna(ech_max) else f"{int(ech_max)}h"
                print(f"   ⏳ {nom} encore partiel ({ech_txt}/{nominal}h), ignoré.")
                continue

            nouveaux[nom] = {'df': df, 'colors': colors, 'info': match.group(1)}
            print(f"   ✅ {nom} récupéré.")
        except Exception as e:
            print(f"   ⚠️ {nom} : connexion échouée ({e}).")

    if not nouveaux:
        print("   Aucun nouveau modèle récupéré pour ce rattrapage.")
        return

    combined = {}
    for sheet, nom in NOM_PAR_SHEET.items():
        existant = lire_feuille_modele(fichier_cible, sheet)
        if existant is not None:
            combined[nom] = existant
    for nom, data in nouveaux.items():
        if nom not in combined:
            combined[nom] = data

    noms_ordonnes = [n for n in ORDRE_MODELES if n in combined]
    construire_et_sauver(
        {n: combined[n]['df'] for n in noms_ordonnes},
        {n: combined[n]['colors'] for n in noms_ordonnes},
        {n: combined[n]['info'] for n in noms_ordonnes},
        fichier_cible
    )


if __name__ == "__main__":
    router_et_construire(extraire_modeles())
    tenter_rattrapage()